import os
import pandas as pd
import time
import logging
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from tqdm import tqdm

# 设置日志
logging.basicConfig(filename='file_info_log.log', level=logging.INFO)


def get_file_info(path, root_path=''):
    file_info = []
    total_files = sum([len(files) for _, _, files in os.walk(path)])

    with tqdm(total=total_files, desc="处理文件", unit="file") as pbar:
        for root, dirs, files in os.walk(path):
            for file in files:
                file_path = os.path.join(root, file)

                if not os.path.exists(file_path):
                    logging.warning(f"文件不存在：{file_path}")
                    continue

                size = os.path.getsize(file_path) / 1024 / 1024
                _, ext = os.path.splitext(file)
                creation_time = os.path.getctime(file_path)
                modification_time = os.path.getmtime(file_path)

                file_info.append({
                    '文件名': file,
                    '文件大小(MB)': size,
                    '文件类型': ext,
                    '文件夹路径': os.path.relpath(root, root_path),
                    '创建时间': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(creation_time)),
                    '最后修改时间': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(modification_time)),
                    '最后修改人': get_user_name()
                })

                pbar.update(1)

    return file_info


def save_to_excel(file_info, output_path):
    df = pd.DataFrame(file_info)
    max_rows_per_file = 1000000

    num_files = len(df) // max_rows_per_file + (1 if len(df) % max_rows_per_file != 0 else 0)

    for file_num in range(num_files):
        start_row = file_num * max_rows_per_file
        end_row = (file_num + 1) * max_rows_per_file
        current_df = df.iloc[start_row:end_row]

        if num_files > 1:
            current_output_path = output_path.replace('.xlsx', f'_part{file_num + 1}.xlsx')
        else:
            current_output_path = output_path

        writer = pd.ExcelWriter(current_output_path, engine='openpyxl')
        current_df.to_excel(writer, index=False)
        wb = writer.book
        ws = wb.active

        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name='微软雅黑')

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length

        writer.close()


def get_user_name():
    import getpass
    return getpass.getuser()


if __name__ == '__main__':
    source_path = input("请输入文件夹路径：")
    start_time = time.time()

    folder_name = os.path.basename(source_path)
    date_time = time.strftime('%Y%m%d-%H%M%S', time.localtime())
    base_output_folder = os.environ.get('BASE_OUTPUT_FOLDER', '/path/to/output/folder')  # 使用环境变量或自定义
    base_output_path = os.path.join(base_output_folder, f'{folder_name}_{date_time}.xlsx')

    file_info = get_file_info(source_path, source_path)
    save_to_excel(file_info, base_output_path)

    end_time = time.time()
    duration = end_time - start_time

    print(f"文件信息已保存到：{base_output_path}")
    print(f"本次运行耗时：{duration:.2f}秒")
